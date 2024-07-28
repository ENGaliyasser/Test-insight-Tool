
import importlib
from lxml import etree
from bs4 import BeautifulSoup
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from selenium import webdriver
import pandas as pd
import shutil
import datetime

from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.select import Select
from selenium import webdriver
from PyQt5.QtGui import QTextCursor
import PyQt5.QtWidgets
import os
from pyexcel.cookbook import merge_all_to_a_book
import glob
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.alert import Alert
import numpy as np
from tabulate import tabulate
import pandas as pd
import traceback
from gui import Ui_MainWindow
from PyQt5 import QtCore, QtGui, QtWidgets
import sys
import os
import time
from PyQt5.QtWidgets import QMessageBox, QFileDialog
import threading
from openpyxl import load_workbook , Workbook
import sys
import pandas as pd
import sys
import pandas as pd
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from gui import Ui_MainWindow  # Importing the UI class from the converted gui.py file
import sys
import pandas as pd
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QVBoxLayout
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from gui import Ui_MainWindow  # Importing the UI class from the converted gui.py file

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QTextCursor
from PyQt5 import uic
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


class Back_End_Class(QtWidgets.QWidget, Ui_MainWindow):

    def __init__(self):
        QtWidgets.QWidget.__init__(self)
        self.thread = {}
        super().__init__()
        self.setupUi(MainWindow)
        self.file_path = None
        self.dir_name = None
        self.browse_btn.clicked.connect(self.browse_folder)
        self.extract.clicked.connect(self.extract_data)

    def browse_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder:
            self.browse_line.setText(folder)

    def extract_data(self):
        folder = self.browse_line.text()
        if not os.path.isdir(folder):
            self.textBrowser.append("Invalid folder path.")
            return

        html_files = self.get_html_files(folder)
        if not html_files:
            self.textBrowser.append("No HTML files found in the selected folder.")
            return

        wb, ws = self.create_workbook_with_headers()
        summary = []
        test_no = 1

        driver = webdriver.Chrome()

        for html_file in html_files:
            file_path = os.path.join(folder, html_file)
            driver.get(f"file:///{file_path}")
            rows = driver.find_elements(By.XPATH, '//table/tbody/tr')

            self.append_test_steps(ws, rows, test_no)
            overall_result = self.extract_overall_result(rows)
            self.append_overall_result(ws, overall_result, test_no)

            summary.append(f"Test {test_no}: {os.path.basename(html_file)} - {overall_result}")
            test_no += 1

        self.adjust_column_widths(ws)
        driver.quit()
        self.save_workbook(wb, folder)

        self.display_summary(summary)

    def get_html_files(self, folder):
        """Return a list of HTML files in the given folder."""
        return [f for f in os.listdir(folder) if f.endswith('.html')]

    def create_workbook_with_headers(self):
        """Create a workbook with headers and return the workbook and worksheet."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"
        headers = ["Test No.", "Test Step", "Description", "Expected Result", "Obtained Result","Step Result", "Overall Test Result"]
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"].font = Font(bold=True)

        return wb, ws

    def append_test_steps(self, ws, rows, test_no):
        """Append the test steps to the worksheet."""
        for row in rows[:-1]:  # Exclude the last row for test result
            cols = row.find_elements(By.TAG_NAME, 'td')
            # ws.append([test_no] + [col.text for col in cols])
            ws.append([test_no] + [col.text for col in cols[:]] + [""])

    def extract_overall_result(self, rows):
        """Extract the overall result from the last row."""
        return rows[-1].find_element(By.TAG_NAME, 'p').text.split(':')[1].strip()

    def append_overall_result(self, ws, overall_result, test_no):
        """Append the overall result to the worksheet."""
        # ws.append([test_no, "", "", "", "", overall_result])
        last_row = ws.max_row
        last_col = ws.max_column

        ws.cell(row=last_row , column=7, value=overall_result)


    def adjust_column_widths(self, ws):
        """Adjust the column widths based on the content."""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    def save_workbook(self, wb, folder):
        """Save the workbook to the specified folder."""
        wb.save(os.path.join(folder, "result.xlsx"))

    def display_summary(self, summary):
        """Display the summary of the results in the text browser."""
        self.textBrowser.clear()
        self.textBrowser.append(f"Number of tests: {len(summary)}")
        for line in summary:
            self.textBrowser.append(line)
            self.textBrowser.moveCursor(QTextCursor.End)



if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Back_End_Class()
    MainWindow.show()
    sys.exit(app.exec_())
