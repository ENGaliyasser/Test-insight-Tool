import sys  # Required for accessing command-line arguments and exiting the application
from PyQt5.QtCore import QThread, pyqtSignal, QObject, QMutex, QMutexLocker  # PyQt5 modules for threading and signals
from PyQt5.QtGui import QTextCursor  # PyQt5 module for handling text cursor in text widgets
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox  # PyQt5 modules for creating the GUI
import os  # Required for file and directory operations
import time  # Required for sleep operations in threading
from selenium import webdriver  # Required for web scraping with Selenium
from selenium.webdriver.chrome.options import Options  # Required for headless browser options
from selenium.webdriver.common.by import By  # Required for locating elements in the web page
from openpyxl import Workbook  # Required for creating and manipulating Excel workbooks
from openpyxl.utils import get_column_letter  # Required for column letter conversion in Excel
from openpyxl.styles import Font  # Required for styling Excel cells
from gui import Ui_MainWindow  # Importing the UI design

# Global variable to track the progress value
progress_value = 0


class TaskThread(QThread):
    """
    QThread subclass for performing file extraction tasks in a separate thread.
    """
    progress = pyqtSignal(int)
    update_summary = pyqtSignal(list)
    finished = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(self, index, folder=None, parent=None):
        """
        Initialize the thread with the given index and folder.

        :param index: Task index to determine the type of task.
        :param folder: Folder containing HTML files.
        :param parent: Parent QObject.
        """
        super().__init__(parent)
        self.index = index
        self.folder = folder
        self.total_files = 0

    def run(self):
        """
        Main function for the thread. Calls the appropriate function based on the task index.
        """
        try:
            if self.index == 1:  # File Extraction
                self.handle_file_extraction()

        except Exception as e:
            self.error.emit(f"An unexpected error occurred: {str(e)}")
            self.finished.emit()

    def handle_file_extraction(self):
        """
        Handles the file extraction process, including initial validation, setting up workbook, and processing HTML files.
        """
        if not self.validate_folder_path():
            return

        html_files = self.get_html_files(self.folder)
        self.total_files = len(html_files)

        if not self.check_html_files(html_files):
            return

        wb, ws = self.create_workbook_with_headers()
        summary = []
        driver = self.setup_webdriver()

        try:
            self.process_html_files(html_files, ws, summary, driver)
        finally:
            self.cleanup_driver(driver)
            self.save_workbook(wb, self.folder)
            self.finished.emit()

    def validate_folder_path(self):
        """
        Validate the folder path. Emits an error signal if the folder path is not set.

        :return: Boolean indicating whether the folder path is valid.
        """
        if not self.folder:
            self.error.emit("Folder path is not set.")
            self.finished.emit()
            return False
        return True

    def check_html_files(self, html_files):
        """
        Check if there are HTML files in the selected folder. Emits an update signal if no HTML files are found.

        :param html_files: List of HTML files.
        :return: Boolean indicating whether HTML files are present.
        """
        if not html_files:
            self.update_summary.emit(["No HTML files found in the selected folder."])
            self.finished.emit()
            return False
        return True

    def setup_webdriver(self):
        """
        Setup and return a headless Chrome WebDriver.

        :return: Configured WebDriver instance.
        """
        options = Options()
        options.add_argument('--headless')
        return webdriver.Chrome(options=options)

    def process_html_files(self, html_files, ws, summary, driver):
        """
        Process a list of HTML files, updating the workbook and summary.

        :param html_files: List of HTML files to process.
        :param ws: The worksheet to append data to.
        :param summary: List to store summary of test results.
        :param driver: The WebDriver instance.
        """
        test_no = 1
        for i, html_file in enumerate(html_files):
            self.process_single_html_file(html_file, ws, summary, test_no, driver)
            progress_value = int(((i + 1) / self.total_files) * 100)
            self.progress.emit(progress_value)
            self.update_summary.emit(summary)
            test_no += 1

    def process_single_html_file(self, html_file, ws, summary, test_no, driver):
        """
        Process a single HTML file and append results to the workbook and summary.

        :param html_file: The HTML file to process.
        :param ws: The worksheet to append data to.
        :param summary: List to store summary of test results.
        :param test_no: Current test number.
        :param driver: The WebDriver instance.
        """
        try:
            file_path = os.path.join(self.folder, html_file)
            driver.get(f"file:///{file_path}")
            rows = driver.find_elements(By.XPATH, '//table/tbody/tr')

            self.append_test_steps(ws, rows, test_no)
            overall_result = self.extract_overall_result(rows)
            self.append_overall_result(ws, overall_result, test_no)

            summary.append(f"Test {test_no}: {os.path.basename(html_file)} - {overall_result}")

        except Exception as e:
            self.error.emit(f"Error processing file {html_file}: {str(e)}")

    def get_html_files(self, folder):
        """
        Get a list of HTML files in the given folder.

        :param folder: The directory path to search for HTML files.
        :return: List of HTML file names.
        """
        return [f for f in os.listdir(folder) if f.endswith('.html')]

    def create_workbook_with_headers(self):
        """
        Create an Excel workbook with headers for test results.

        :return: Tuple containing workbook and worksheet.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"
        headers = ["Test No.", "Test Step", "Description", "Expected Result", "Obtained Result", "Step Result",
                   "Overall Test Result"]
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"].font = Font(bold=True)

        return wb, ws

    def append_test_steps(self, ws, rows, test_no):
        """
        Append test steps to the worksheet.

        :param ws: The worksheet to append data to.
        :param rows: List of rows from the HTML table.
        :param test_no: Current test number.
        """
        for row in rows[:-1]:
            cols = row.find_elements(By.TAG_NAME, 'td')
            ws.append([test_no] + [col.text for col in cols[:]] + [""])

    def extract_overall_result(self, rows):
        """
        Extract the overall result from the last row of the table.

        :param rows: List of rows from the HTML table.
        :return: The overall result string.
        """
        return rows[-1].find_element(By.TAG_NAME, 'p').text.split(':')[1].strip()

    def append_overall_result(self, ws, overall_result, test_no):
        """
        Append the overall result to the last column of the last row in the worksheet.

        :param ws: The worksheet to append data to.
        :param overall_result: The overall result string.
        :param test_no: Current test number.
        """
        last_row = ws.max_row
        ws.cell(row=last_row, column=7, value=overall_result)

    def cleanup_driver(self, driver):
        """
        Cleanup and quit the WebDriver instance.

        :param driver: The WebDriver instance to quit.
        """
        driver.quit()

    def save_workbook(self, wb, folder):
        """
        Save the workbook to the given folder with the filename 'result.xlsx'.

        :param wb: The workbook to save.
        :param folder: The directory path to save the workbook in.
        """
        wb.save(os.path.join(folder, "result.xlsx"))


class Back_End_Class(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.file_path = None
        self.dir_name = None
        self.browse_btn.clicked.connect(self.browse_folder)
        self.extract.clicked.connect(self.start_extraction)
        self.task_thread = None

    def browse_folder(self):
        """
        Open a dialog to select a folder and display the selected folder path in the line edit.
        """
        folder = QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder:
            if not os.listdir(folder):  # Check if the folder is empty
                self.show_error_message("The selected folder is empty.")
            else:
                self.browse_line.setText(folder)

    def start_extraction(self):
        """
        Start the extraction process by initializing and starting the progress and extraction threads.
        """
        self.extract.setEnabled(False)
        folder = self.browse_line.text()
        if not os.path.isdir(folder):
            self.show_error_message("Wrong Path.")
            return

        html_files = self.get_html_files(folder)
        if not html_files:
            self.show_error_message("No HTML files found in the selected folder.")
            return



        # Start the extraction thread
        self.task_thread = TaskThread(index=1, folder=folder)
        self.task_thread.update_summary.connect(self.display_summary)
        self.task_thread.progress.connect(self.update_progress_bar)
        self.task_thread.finished.connect(self.on_extraction_finished)
        self.task_thread.error.connect(self.show_error_message)
        self.task_thread.start()

    def update_progress_bar(self, value):
        """
        Update the progress bar with the given value.
        :param value: Progress value to set.
        """
        self.progressBar.setValue(value)

    def on_extraction_finished(self):
        """
        Display a message when the extraction and processing are completed.
        """
        self.textBrowser.append("Extraction and processing completed.")
        self.extract.setEnabled(True)


    def display_summary(self, summary):
        """
        Display the summary of tests in the text browser.
        :param summary: List of summary lines to display.
        """
        self.textBrowser.clear()
        self.textBrowser.append(f"Number of tests: {self.task_thread.total_files}")
        for line in summary:
            self.textBrowser.append(line)
            self.textBrowser.moveCursor(QTextCursor.End)

    def get_html_files(self, folder):
        """
        Get a list of HTML files in the given folder.
        :param folder: The directory path to search for HTML files.
        :return: List of HTML file names.
        """
        return [f for f in os.listdir(folder) if f.endswith('.html')]

    def show_error_message(self, message):
        """
        Show an error message box with the given message.
        :param message: The error message to display.
        """
        QMessageBox.warning(self, "Error", message)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = Back_End_Class()
    window.show()
    sys.exit(app.exec_())
