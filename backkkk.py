import sys

from PyQt5.QtCore import QThread, pyqtSignal, QObject, QMutex, QMutexLocker
from PyQt5.QtGui import QTextCursor
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QVBoxLayout
import os
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from gui import Ui_MainWindow
progress_value = 0

class SharedData(QObject):
    def __init__(self):
        super().__init__()
        self._value = None
        self._lock = QMutex()

    def set_value(self, value):
        with QMutexLocker(self._lock):
            self._value = value

    def get_value(self):
        with QMutexLocker(self._lock):
            return self._value
class TaskThread(QThread):
    progress = pyqtSignal(int)
    update_summary = pyqtSignal(list)
    finished = pyqtSignal()

    def __init__(self, index, folder=None, parent=None):
        super().__init__(parent)
        self.index = index
        self.folder = folder
        self.total_files = 0
        # self.progress_value = 0

    def run(self):
        if self.index == 1:  # Progress Bar Update
            """Monitor the progress value and emit changes."""
            global progress_value

            progress_value = 0
            last_value = -1
            while True:
                if progress_value != last_value:
                    self.progress.emit(progress_value)
                    last_value = progress_value

                if progress_value >= 100:
                    break

                 # Check every 0.01 seconds

        elif self.index == 2:  # File Extraction
            if not self.folder:
                self.finished.emit()
                return

            html_files = self.get_html_files(self.folder)
            self.total_files = len(html_files)
            if not html_files:
                self.update_summary.emit(["No HTML files found in the selected folder."])
                self.finished.emit()
                return

            wb, ws = self.create_workbook_with_headers()
            summary = []
            test_no = 1
            options = Options()  # Initalizing options variable to contain chrome driver options.
            options.add_argument('--headless=new')  # Headless option to run chrome in silent mode.
            driver = webdriver.Chrome(options)

            for i, html_file in enumerate(html_files):
                file_path = os.path.join(self.folder, html_file)
                driver.get(f"file:///{file_path}")
                rows = driver.find_elements(By.XPATH, '//table/tbody/tr')

                self.append_test_steps(ws, rows, test_no)
                overall_result = self.extract_overall_result(rows)
                self.append_overall_result(ws, overall_result, test_no)

                summary.append(f"Test {test_no}: {os.path.basename(html_file)} - {overall_result}")

                # Update progress and summary

                progress_value = int(((i + 1) / self.total_files) * 100)
                # self.progress.emit(progress_value)
                self.update_summary.emit(summary)

                test_no += 1

            self.adjust_column_widths(ws)
            driver.quit()
            self.save_workbook(wb, self.folder)
            self.finished.emit()

    def get_html_files(self, folder):
        return [f for f in os.listdir(folder) if f.endswith('.html')]

    def create_workbook_with_headers(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"
        headers = ["Test No.", "Test Step", "Description", "Expected Result", "Obtained Result", "Step Result", "Overall Test Result"]
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"].font = Font(bold=True)

        return wb, ws

    def append_test_steps(self, ws, rows, test_no):
        for row in rows[:-1]:
            cols = row.find_elements(By.TAG_NAME, 'td')
            ws.append([test_no] + [col.text for col in cols[:]] + [""])

    def extract_overall_result(self, rows):
        return rows[-1].find_element(By.TAG_NAME, 'p').text.split(':')[1].strip()

    def append_overall_result(self, ws, overall_result, test_no):
        last_row = ws.max_row
        ws.cell(row=last_row, column=7, value=overall_result)

    def adjust_column_widths(self, ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    def save_workbook(self, wb, folder):
        wb.save(os.path.join(folder, "result.xlsx"))


class Back_End_Class(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.file_path = None
        self.dir_name = None
        self.browse_btn.clicked.connect(self.browse_folder)
        self.extract.clicked.connect(self.start_extraction)
        self.task_thread = None

    def browse_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder:
            self.browse_line.setText(folder)

    def start_extraction(self):
        folder = self.browse_line.text()
        if not os.path.isdir(folder):
            self.textBrowser.append("Invalid folder path.")
            return

        html_files = self.get_html_files(folder)
        if not html_files:
            self.textBrowser.append("No HTML files found in the selected folder.")
            return
        # Start the extraction thread
        self.progress_thread = TaskThread(index=1, folder=folder)
        self.progress_thread.progress.connect(self.update_progress_bar)
        self.progress_thread.start()
        # Start the extraction thread
        self.task_thread = TaskThread(index=2, folder=folder)
        # self.task_thread.progress.connect(self.update_progress_bar)
        self.task_thread.update_summary.connect(self.display_summary)
        self.task_thread.finished.connect(self.on_extraction_finished)
        self.task_thread.start()

    def update_progress_bar(self, value):
        self.progressBar.setValue(value)

    def on_extraction_finished(self):
        # Actions to perform when extraction thread finishes
        self.textBrowser.append("Extraction and processing completed.")

    def display_summary(self, summary):
        self.textBrowser.clear()
        self.textBrowser.append(f"Number of tests: {self.task_thread.total_files}")
        for line in summary:
            self.textBrowser.append(line)
            self.textBrowser.moveCursor(QTextCursor.End)

    def get_html_files(self, folder):
        return [f for f in os.listdir(folder) if f.endswith('.html')]


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = Back_End_Class()
    window.show()
    sys.exit(app.exec_())
